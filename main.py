import logging
import openpyxl
from openpyxl import Workbook,load_workbook
from script import populate

list_1 = list()
list_2 = list()
list_3 = list()
list_4 = list()

path_1 = r'E:\Python\LMS\Classics.xlsx'
path_2 = r'E:\Python\LMS\Adventure.xlsx'
path_3 = r'E:\Python\LMS\Horror.xlsx'
path_4 = r'E:\Python\LMS\Science Fiction.xlsx'

wb_1 = load_workbook(path_1)
ws_1 = wb_1.active
wb_2 = load_workbook(path_2)
ws_2 = wb_2.active
wb_3 = load_workbook(path_3)
ws_3 = wb_3.active
wb_4 = load_workbook(path_4)
ws_4 = wb_4.active



def getDate():
    import datetime
    now=datetime.datetime.now
    #print("Date: ",now().date())
    return str(now().date())

def getTime():
    import datetime
    now=datetime.datetime.now
    #print("Time: ",now().time())
    return str(now().time())



logging.basicConfig(format='%(message)s', level=logging.INFO)
logging.info('Admin logged in')


class User():
    def __init__(self):
        self.Basic_User = Basic_User
        self._Librarian_User = Librarian_User
        

        

class Book(User):
    def __init__(self, name, ISBN, author):
        self.name = name
        self.ISBN = ISBN
        self.author = author
    def borrow_book():
        
        Name=input("Enter your name: ")
        Genre_name = int(input("Enter the number of the genre whose book you wish to borrow: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction\n"))
        
        if Genre_name == 1:
            print("The following books are available in Classics: ")
            for row in range(2,7):
                print("Serial Number: ", ws_1['A' + str(row)].value, "Book Name: ", ws_1['B' + str(row)].value, "ISBN: ", ws_1['C' + str(row)].value, "Author: ", ws_1['D' + str(row)].value)

            Choice_1 = int(input("Enter serial number of the book you wish to borrow: "))
            logging.info('Book number : %d', Choice_1)
            logging.info('Has been borrowed by: %s', Name)
                           
                
                
        elif Genre_name == 2:
            
            print("The following books are available in Adventure: ")
            for row in range(2,7):
                print("Serial Number: ", ws_2['A' + str(row)].value, "Book Name: ", ws_2['B' + str(row)].value, "ISBN: ", ws_2['C' + str(row)].value, "Author: ", ws_2['D' + str(row)].value)

            Choice_2 = int(input("Enter serial number of the book you wish to borrow: "))
            logging.info('Book number : %d', Choice_2)
            logging.info('Has been borrowed by: %s', Name)
                           
        elif Genre_name == 3:
            
            print("The following books are available in Horror: ")
            for row in range(2,7):
                print("Serial Number: ", ws_3['A' + str(row)].value, "Book Name: ", ws_3['B' + str(row)].value, "ISBN: ", ws_3['C' + str(row)].value, "Author: ", ws_3['D' + str(row)].value)

            Choice_3 = int(input("Enter serial number of the book you wish to borrow: "))
            logging.info('Book number : %d', Choice_3)
            logging.info('Has been borrowed by: %s', Name)
                           
        elif Genre_name == 4:
            
            print("The following books are available in Science Fiction: ")
            for row in range(2,7):
                print("Serial Number: ", ws_4['A' + str(row)].value, "Book Name: ", ws_4['B' + str(row)].value, "ISBN: ", ws_4['C' + str(row)].value, "Author: ", ws_4['D' + str(row)].value)

            Choice_4 = int(input("Enter serial number of the book you wish to borrow: "))
            logging.info('Book number : %d', Choice_4)
            logging.info('Has been borrowed by: %s', Name)
                           
        else:
            print("Error")
        

            
    def return_book():
        
        Name=input("Enter your name: ")
        Genre_name = int(input("Enter the number of the genre whose book you wish to return: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction"))
        
        if Genre_name == 1:
            print("The following books are available in Classics: ")
            for row in range(2,7):
                print("Serial Number: ", ws_1['A' + str(row)].value, "Book Name: ", ws_1['B' + str(row)].value, "ISBN: ", ws_1['C' + str(row)].value, "Author: ", ws_1['D' + str(row)].value)

            Choice_5 = int(input("Enter serial number of the book you wish to return: "))
            logging.info('Book number : %d', Choice_5)
            logging.info('Has been returned by: %s', Name)
                           
                
                
        elif Genre_name == 2:
            
            print("The following books are available in Adventure: ")
            for row in range(2,7):
                print("Serial Number: ", ws_2['A' + str(row)].value, "Book Name: ", ws_2['B' + str(row)].value, "ISBN: ", ws_2['C' + str(row)].value, "Author: ", ws_2['D' + str(row)].value)

            Choice_6 = int(input("Enter serial number of the book you wish to return: "))
            logging.info('Book number : %d', Choice_6)
            logging.info('Has been returned by: %s', Name)
                           
        elif Genre_name == 3:
            
            print("The following books are available in Horror: ")
            for row in range(2,7):
                print("Serial Number: ", ws_3['A' + str(row)].value, "Book Name: ", ws_3['B' + str(row)].value, "ISBN: ", ws_3['C' + str(row)].value, "Author: ", ws_3['D' + str(row)].value)

            Choice_7 = int(input("Enter serial number of the book you wish to return: "))
            logging.info('Book number : %d', Choice_7)
            logging.info('Has been returned by: %s', Name)
                           
        elif Genre_name == 4:
            
            print("The following books are available in Science Fiction: ")
            for row in range(2,7):
                print("Serial Number: ", ws_4['A' + str(row)].value, "Book Name: ", ws_4['B' + str(row)].value, "ISBN: ", ws_4['C' + str(row)].value, "Author: ", ws_4['D' + str(row)].value)

            Choice_8 = int(input("Enter serial number of the book you wish to return: "))
            logging.info('Book number : %d', Choice_8)
            logging.info('Has been returned by: %s', Name)
                           
        else:
            print("Error")
        



        
        
    def reserve_book():
        Name=input("Enter your name: ")
        Genre_name = int(input("Enter the number of the genre whose book you wish to reserve: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction\n"))
        
        if Genre_name == 1:
            print("The following books are available in Classics: ")
            for row in range(2,7):
                print("Serial Number: ", ws_1['A' + str(row)].value, "Book Name: ", ws_1['B' + str(row)].value, "ISBN: ", ws_1['C' + str(row)].value, "Author: ", ws_1['D' + str(row)].value)

            Choice_9 = int(input("Enter serial number of the book you wish to reserve: "))
            logging.info('Book number : %d', Choice_9)
            logging.info('Has been reserved by: %s', Name)
                           
                
                
        elif Genre_name == 2:
            
            print("The following books are available in Adventure: ")
            for row in range(2,7):
                print("Serial Number: ", ws_2['A' + str(row)].value, "Book Name: ", ws_2['B' + str(row)].value, "ISBN: ", ws_2['C' + str(row)].value, "Author: ", ws_2['D' + str(row)].value)

            Choice_10 = int(input("Enter serial number of the book you wish to reserve: "))
            logging.info('Book number : %d', Choice_10)
            logging.info('Has been reserved by: %s', Name)
                           
        elif Genre_name == 3:
            
            print("The following books are available in Horror: ")
            for row in range(2,7):
                print("Serial Number: ", ws_3['A' + str(row)].value, "Book Name: ", ws_3['B' + str(row)].value, "ISBN: ", ws_3['C' + str(row)].value, "Author: ", ws_3['D' + str(row)].value)

            Choice_11 = int(input("Enter serial number of the book you wish to reserve: "))
            logging.info('Book number : %d', Choice_11)
            logging.info('Has been reserved by: %s', Name)
                           
        elif Genre_name == 4:
            
            print("The following books are available in Science Fiction: ")
            for row in range(2,7):
                print("Serial Number: ", ws_4['A' + str(row)].value, "Book Name: ", ws_4['B' + str(row)].value, "ISBN: ", ws_4['C' + str(row)].value, "Author: ", ws_4['D' + str(row)].value)

            Choice_12 = int(input("Enter serial number of the book you wish to reserve: "))
            logging.info('Book number : %d', Choice_12)
            logging.info('Has been reserved by: %s', Name)
                           
        else:
            print("Error")
        



class Shelf():
    def __init__(self, genre):
        self.genre = genre
    def show_catalog():
        Genre_name = int(input("Enter the number of the genre whose books you wish to view: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction\n"))
        
        if Genre_name == 1:
            
            print("The following books are available in Classics: ")
            for row in range(2,7):
                print("Serial Number: ", ws_1['A' + str(row)].value, "Book Name: ", ws_1['B' + str(row)].value, "ISBN: ", ws_1['C' + str(row)].value, "Author: ", ws_1['D' + str(row)].value)

        elif Genre_name == 2:
            print("The following books are available in Adventure: ")
            for row in range(2,7):
                print("Serial Number: ", ws_2['A' + str(row)].value, "Book Name: ", ws_2['B' + str(row)].value, "ISBN: ", ws_2['C' + str(row)].value, "Author: ", ws_2['D' + str(row)].value)                

        elif Genre_name == 3:
            print("The following books are available in Horror: ")
            for row in range(2,7):
                print("Serial Number: ", ws_3['A' + str(row)].value, "Book Name: ", ws_3['B' + str(row)].value, "ISBN: ", ws_3['C' + str(row)].value, "Author: ", ws_3['D' + str(row)].value)
        elif Genre_name == 4:
            print("The following books are available in Science Fiction: ")
            for row in range(2,7):
                print("Serial Number: ", ws_4['A' + str(row)].value, "Book Name: ", ws_4['B' + str(row)].value, "ISBN: ", ws_4['C' + str(row)].value, "Author: ", ws_4['D' + str(row)].value)


    def get_books_count():
        Genre_name = int(input("Enter the number of the genre whose number of books you wish to count: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction\n"))
        
        if Genre_name == 1:
            
            for row in range(2,7):
                list_1.append(ws_1['A' + str(row)].value)
                
            print("Number of books in Classics genre is: ", len(list_1))

        elif Genre_name == 2:
            for row in range(2,7):
                list_2.append(ws_2['A' + str(row)].value)
                
            print("Number of books in Adventure genre is: ", len(list_2))

        elif Genre_name == 3:
            for row in range(2,7):
                list_3.append(ws_3['A' + str(row)].value)
                
            print("Number of books in Horror genre is: ", len(list_3))

        elif Genre_name == 4:
            for row in range(2,7):
                list_4.append(ws_4['A' + str(row)].value)
                
            print("Number of books in Science Fiction genre is: ", len(list_4))
            
        
                
        
   


class Shelf_Lib(User):
    def __init__(self):
        pass

    def add_book():
        Genre_name = int(input("Enter the number of the genre of the book which you wish to add: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction\n"))
        
        if Genre_name == 1:
            
            n = int(input("Enter number of books you wish to add in genre 1. Classics : "))
            for i in range(n):
                    Book_name = input("Enter the name of the book: ")
                    Book_ISBN = int(input("Enter the ISBN of the book: "))
                    Book_Author = input("Enter name of the author of the book: ")
                    data = ((6+i), (Book_name), (Book_ISBN), (Book_Author))
                    ws_1.append(data)
                    wb_1.save(path_1)
                    
        elif Genre_name == 2:
            n = int(input("Enter number of books you wish to add in genre 2. Adventure : "))
            for i in range(n):
                    Book_name = input("Enter the name of the book: ")
                    Book_ISBN = int(input("Enter the ISBN of the book: "))
                    Book_Author = input("Enter name of the author of the book: ")
                    data = ((6+i), (Book_name), (Book_ISBN), (Book_Author))
                    ws_2.append(data)
                    wb_2.save(path_2)
            
        elif Genre_name == 3:
            n = int(input("Enter number of books you wish to add in genre 3. Horror : "))
            for i in range(n):
                    Book_name = input("Enter the name of the book: ")
                    Book_ISBN = int(input("Enter the ISBN of the book: "))
                    Book_Author = input("Enter name of the author of the book: ")
                    data = ((6+i), (Book_name), (Book_ISBN), (Book_Author))
                    ws_3.append(data)
                    wb_3.save(path_3)
                    
        elif Genre_name == 4:
            n = int(input("Enter number of books you wish to add in genre 4. Science and Fiction : "))
            for i in range(n):
                    Book_name = input("Enter the name of the book: ")
                    Book_ISBN = int(input("Enter the ISBN of the book: "))
                    Book_Author = input("Enter name of the author of the book: ")
                    data = ((6+i), (Book_name), (Book_ISBN), (Book_Author))
                    ws_4.append(data)
                    wb_4.save(path_4)

    def remove_book():
        Genre_name = int(input("Enter the number of the genre of the book which you wish to remove: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction"))
        if Genre_name == 1:
            
            print("The following books are available in Classics: ")
            for row in range(2,7):
                print("Serial Number: ", ws_1['A' + str(row)].value, "Book Name: ", ws_1['B' + str(row)].value, "ISBN: ", ws_1['C' + str(row)].value, "Author: ", ws_1['D' + str(row)].value)

            rm = int(input("Enter serial number of the book you wish to remove: "))
            ws_1.delete_rows(idx = rm + 1)
            wb_1.save(path_1)

    
            
        elif Genre_name == 2:
            
            print("The following books are available in Adventure: ")
            for row in range(2,7):
                print("Serial Number: ", ws_2['A' + str(row)].value, "Book Name: ", ws_2['B' + str(row)].value, "ISBN: ", ws_2['C' + str(row)].value, "Author: ", ws_2['D' + str(row)].value)                

            rm = int(input("Enter serial number of the book you wish to remove: "))
            ws_2.delete_rows(idx = rm + 1)
            wb_2.save(path_2)

        elif Genre_name == 3:
            
            print("The following books are available in Horror: ")
            for row in range(2,7):
                print("Serial Number: ", ws_3['A' + str(row)].value, "Book Name: ", ws_3['B' + str(row)].value, "ISBN: ", ws_3['C' + str(row)].value, "Author: ", ws_3['D' + str(row)].value)

            rm = int(input("Enter serial number of the book you wish to remove: "))
            ws_3.delete_rows(idx = rm + 1)
            wb_3.save(path_3)
            
        elif Genre_name == 4:

            print("The following books are available in Science Fiction: ")
            for row in range(2,7):
                print("Serial Number: ", ws_4['A' + str(row)].value, "Book Name: ", ws_4['B' + str(row)].value, "ISBN: ", ws_4['C' + str(row)].value, "Author: ", ws_4['D' + str(row)].value)

            rm = int(input("Enter serial number of the book you wish to remove: "))
            ws_4.delete_rows(idx = rm + 1)
            wb_4.save(path_4)
        

    def populate_book():
        genre_name = int(input("Enter the number of the genre of the book which you wish to populate: \n 1. Classics \n 2. Adventure \n 3. Horror \n 4. Science Fiction\n"))
        populate(genre_name)

        
a=int(input('''Welcome to the Library Management System. Are you a basic user(enter 1) or a librarian user (enter 0)?  '''))
if a == 1:
    ch = int(input('''What brings you here today? Select the corresponding number:
                        1. Borrow Book
                        2. Return Book
                        3. Reserve Book
                        4. View catalog
                        5. Get book count\n'''))

    if ch == 1:
        Book.borrow_book()
    elif ch == 2:
        Book.return_book()
    elif ch == 3:
        Book.reserve_book()
    elif ch == 4:
        Shelf.show_catalog()
    elif ch == 5:
        Shelf.get_books_count()
    else:
        print("Invalid choice")
elif a == 0:
    ch = int(input('''What brings you here today? Select the corresponding number:
                        1. View Catalog
                        2. Get Book Count
                        3. Add Book
                        4. Remove Book
                        5. Populate Book\n'''))
                        
    if ch == 1:
        Shelf.show_catalog()
    elif ch == 2:
        Shelf.get_books_count()
    elif ch == 3:
        Shelf_Lib.add_book()
    elif ch == 4:
        Shelf_Lib.remove_book()
    elif ch == 5:
        Shelf_Lib.populate_book()  
    else:
        print("Invalid choice")
    
else:
    print("Invalid choice")




    


        
